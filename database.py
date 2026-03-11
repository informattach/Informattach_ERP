import os
from dotenv import load_dotenv
from supabase import create_client, Client
import streamlit as st
from typing import List, Dict, Optional

class DatabaseManager:
    def __init__(self):
        try:
            url: str = st.secrets["SUPABASE_URL"]
            key: str = st.secrets["SUPABASE_KEY"]
        except (FileNotFoundError, KeyError):
            load_dotenv()
            url: str = os.environ.get("SUPABASE_URL")
            key: str = os.environ.get("SUPABASE_KEY")
            
        if not url or not key:
            raise ValueError("SUPABASE_URL ve SUPABASE_KEY bulunamadı!")
        
        self.client: Client = create_client(url, key)

    def create_core_product(self, base_title: str, isku: str, asin: str = None, upc: str = None, requires_expiration: bool = False) -> Dict:
        prod_data = {
            "asin": asin, 
            "upc": upc, 
            "requires_expiration": requires_expiration
        }
        prod_data = {k: v for k, v in prod_data.items() if v is not None} 
        
        new_prod = self.client.table("core_products").insert(prod_data).execute()
        product_id = new_prod.data[0]['id']
        
        self.client.table("product_base_content").insert({
            "product_id": product_id,
            "base_title": base_title
        }).execute()
        
        self.client.table("product_logistics").insert({
            "product_id": product_id
        }).execute()
        
        # ISKU'yu listings tablosuna ekle
        if isku:
            STORE_ID = "197bd215-3bec-4f43-aa40-f2fb4d204eee"
            self.client.table("listings").insert({
                "product_id": product_id,
                "store_id": STORE_ID,
                "channel_sku": isku,
                "is_active": False,
                "needs_sync": False
            }).execute()
        
        return new_prod.data[0]

    def get_product_by_asin(self, asin: str) -> Optional[Dict]:
        response = self.client.table("core_products").select(
            "*, product_base_content(*), product_logistics(*), product_media(*), sources(*), listings(*), product_documents(*)"
        ).eq("asin", asin).execute()
        
        if response.data:
            return response.data[0]
        return None

    def get_all_core_products(self) -> List[Dict]:
        """Ana ürünleri listeleme ekranı için temel bilgilerle getirir."""
        all_products = []
        offset = 0
        limit = 1000  # Supabase default max_rows is 1000
        while True:
            response = self.client.table("core_products").select(
                "id, asin, upc, created_at, product_media(media_url), product_base_content(base_title, updated_at), requires_expiration, "
                "sources(base_cost, supplier_id, updated_at), listings(channel_item_id, listed_price, quantity, channel_sku, category_id, store_id, shipping_profile_id, return_profile_id, payment_profile_id, updated_at), "
                "product_documents(document_type, document_url)"
            ).order("created_at", desc=True).range(offset, offset + limit - 1).execute()
            
            data = response.data
            if data is None:
                break # in case of error
                
            all_products.extend(data)
            if len(data) < limit:
                break
            offset += limit
            
        return all_products
        
    def get_all_categories(self) -> Dict[str, str]:
        all_cats = []
        offset = 0
        limit = 1000
        while True:
            res = self.client.table("marketplace_categories").select("category_id, category_name").range(offset, offset+limit-1).execute()
            if not res.data: break
            all_cats.extend(res.data)
            if len(res.data) < limit: break
            offset += limit
        return {c['category_id']: c['category_name'] for c in all_cats}

    def get_all_stores(self) -> Dict[str, str]:
        res = self.client.table("stores").select("id, store_name").execute()
        return {s['id']: s['store_name'] for s in res.data} if res.data else {}

    def get_all_suppliers(self) -> Dict[str, str]:
        res = self.client.table("suppliers").select("id, name").execute()
        return {s['id']: s['name'] for s in res.data} if res.data else {}

    def import_easync_data(self, df) -> dict:
        """Toplu işlem (Bulk Insert) mimarisi ile sıfır ağ gecikmesi."""
        st.info("🚀 Veriler paketleniyor (Bulk İşlem)...")
        df = df.fillna('')
        
        # --- 1. MEVCUT VERİYİ RAM'E ÇEK (Sayfalama ile 1000 sınırını aşarak) ---
        mp_data = self.client.table('marketplaces').select('id, name, region').execute().data
        mp_cache = {f"{m['name']}_{m['region']}": m['id'] for m in mp_data}
        
        sup_data = self.client.table('suppliers').select('id, name').execute().data
        supplier_cache = {s['name']: s['id'] for s in sup_data}
        
        store_data = self.client.table('stores').select('id, store_name').execute().data
        store_cache = {s['store_name']: s['id'] for s in store_data}
        
        # core_products için 1000 sınırını aşacak şekilde tek tek ID'leri çekelim veya CSV'den ASIN'leri bulup sadece onları IN ile çekelim
        # Daha temizi, Supabase'in sayfalama (limit/offset veya range) özelliğini kullanmak
        source_ids_in_csv = df['Source Product Id'].dropna().astype(str).unique().tolist()
        
        prod_data = []
        chunk_sz = 200
        for i in range(0, len(source_ids_in_csv), chunk_sz):
            chunk_asins = source_ids_in_csv[i:i+chunk_sz]
            res = self.client.table('core_products').select('id, asin').in_('asin', chunk_asins).execute()
            prod_data.extend(res.data)
            
        prod_cache = {p['asin']: p['id'] for p in prod_data} # asin üzerinden ID eşleştir
        
        # Şimdilik listing ve source cache'lerini çok büyük olmaması için sadece DB'de olan product'lar için alalım
        existing_p_ids = [p['id'] for p in prod_data]
        list_data = []
        src_data = []
        if existing_p_ids:
            for i in range(0, len(existing_p_ids), chunk_sz):
                p_chunk = existing_p_ids[i:i+chunk_sz]
                l_res = self.client.table('listings').select('id, product_id, store_id').in_('product_id', p_chunk).execute()
                list_data.extend(l_res.data)
                
                s_res = self.client.table('sources').select('id, product_id, supplier_id').in_('product_id', p_chunk).execute()
                src_data.extend(s_res.data)

        # Cache'de ID'leri tutuyoruz böylece upsert esnasında Primary Key çakışmasını sorunsuz UPDATE edebilelim
        listing_cache = {f"{l['product_id']}_{l['store_id']}": l['id'] for l in list_data}
        source_cache = {f"{s['product_id']}_{s['supplier_id']}": s['id'] for s in src_data}

        # --- PAKET LİSTELERİ (Veritabanına gidecek toplu kargolar) ---
        new_products_payload = []
        new_contents_payload = []
        new_sources_payload = []
        new_listings_payload = []
        
        # --- 2. BİRİNCİ TUR: SADECE YENİ ÜRÜNLERİ TESPİT ET ---
        for index, row in df.iterrows():
            source_id = str(row.get('Source Product Id', '')).strip()
            if not source_id:
                source_id = f"UNKNOWN-{index}"
                
            # Kullanıcının iç SSKU'su (ISKU) AASIN formatıdır
            # Vitrinde (eBay) Target Variant sütunu görünen marka ismidir ama iç sistemdeki kayıtlı ismi AASIN'dir
            true_isku = f"A{source_id}"
            target_variant = str(row.get('Target Variant', '')).strip()
            
            title = str(row.get('Title', 'İsimsiz Ürün'))[:200]

            # Eğer prod_cache'de ASIN yoksa eklenecekler listesine koy
            if source_id not in prod_cache and not any(p['asin'] == source_id for p in new_products_payload):
                new_products_payload.append({'asin': source_id})
                
            # Var olan veya yeni olan TÜM ürünler için ismi de güncelleyelim (Ayrı bir listede biriktiriyoruz)
            # Ancak ID'yi cache'den almalıyız. Eğer yeni eklenecekse, ID'si insert'ten sonra oluşacak.
            # Bu yüzden var olanların başlığını ayrı güncelleyip, yenilerin başlığını insert sonrasına bırakmalıyız.

        # --- 3. TOPLU KAYIT: ÜRÜNLER ---
        if new_products_payload:
            chunk_size = 500
            for i in range(0, len(new_products_payload), chunk_size):
                chunk = new_products_payload[i:i+chunk_size]
                # ASIN kullanarak upsert
                res = self.client.table('core_products').upsert(chunk, on_conflict='asin').execute()
                
                for p in res.data:
                    prod_cache[p['asin']] = p['id']

        # Tüm ID'ler cache'e doldu. Şimdi hepsinin ismini topluca güncelleyelim (İsimsiz Ürün sorununu çözer)
        for index, row in df.iterrows():
            source_id = str(row.get('Source Product Id', '')).strip()
            if not source_id:
                source_id = f"UNKNOWN-{index}"
            
            product_id = prod_cache.get(source_id)
            if product_id:
                title = str(row.get('Title', 'İsimsiz Ürün'))[:200]
                if not any(c['product_id'] == product_id for c in new_contents_payload):
                    new_contents_payload.append({'product_id': product_id, 'base_title': title})
        
        if new_contents_payload:
            chunk_size = 500
            for i in range(0, len(new_contents_payload), chunk_size):
                self.client.table('product_base_content').upsert(new_contents_payload[i:i+chunk_size], on_conflict='product_id').execute()

        # --- 4. İKİNCİ TUR: İLİŞKİLERİ (KAYNAK VE İLAN) PAKETLE ---
        for index, row in df.iterrows():
            source_id = str(row.get('Source Product Id', '')).strip()
            if not source_id:
                source_id = f"UNKNOWN-{index}"
            
            marketplace_sku = str(row.get('Target Variant', '')).strip()
            if not marketplace_sku:
                marketplace_sku = f"A{source_id}"
            
            source_market_raw = str(row.get('Source Market', 'Amazon US'))
            target_market_raw = str(row.get('Target Market', 'eBay US'))
            
            supplier_id = supplier_cache.get(source_market_raw)
            if not supplier_id:
                mp_id = mp_cache.get(f"{source_market_raw.split()[0]}_US", list(mp_cache.values())[0] if mp_cache else None)
                res = self.client.table('suppliers').insert({'marketplace_id': mp_id, 'supplier_type': 'Marketplace_Account', 'name': source_market_raw}).execute()
                supplier_id = res.data[0]['id']
                supplier_cache[source_market_raw] = supplier_id

            store_id = store_cache.get(target_market_raw)
            if not store_id:
                mp_id = mp_cache.get(f"{target_market_raw.split()[0]}_US", list(mp_cache.values())[0] if mp_cache else None)
                res = self.client.table('stores').insert({'marketplace_id': mp_id, 'store_name': target_market_raw}).execute()
                store_id = res.data[0]['id']
                store_cache[target_market_raw] = store_id

            product_id = prod_cache.get(source_id)
            if not product_id:
                continue # Eklendiğine emin olmak için
            
            s_price_raw = str(row.get('Source Price', '0')).replace('$', '').replace(',', '').strip()
            base_cost = float(s_price_raw) if s_price_raw else 0.0
            
            # --- YENİ ERP FİYATLANDIRMA MANTILIĞI ---
            # Easync'in bize dayattığı statik 'Target Price' sütununu çöpe atıyoruz.
            # Sistemi kendi Otonom Pricing Engine'imize bağlıyoruz.
            target_marketplace = 'ebay'
            if 'amazon' in target_market_raw.lower():
                target_marketplace = 'amazon'
            elif 'shopify' in target_market_raw.lower():
                target_marketplace = 'shopify'
                
            from pricing_engine import PricingEngine
            calculated_listed_price = PricingEngine.calculate_final_price(
                source_price=base_cost,
                marketplace=target_marketplace,
                override_marketplace_fee=15.0 # CSV'de kategori olmadığı için şimdilik global %15 fallback
            )
            
            # Easync Quantity yoksa 1 kabul edelim, eğer varsa alalım
            qty_raw = str(row.get('Quantity', '1')).strip()
            try:
                qty = int(qty_raw)
            except:
                qty = 1
            
            src_key = f"{product_id}_{supplier_id}"
            if not any(s['product_id'] == product_id and s['supplier_id'] == supplier_id for s in new_sources_payload):
                s_payload = {
                    'product_id': product_id, 'supplier_id': supplier_id, 
                    'source_code': source_id, 'base_cost': base_cost
                }
                # Eğer db'de varsa id'sini ekle ki upsert doğrudan update yapsın!
                if src_key in source_cache:
                    s_payload['id'] = source_cache[src_key]
                new_sources_payload.append(s_payload)

            lst_key = f"{product_id}_{store_id}"
            if not any(l['product_id'] == product_id and l['store_id'] == store_id for l in new_listings_payload):
                l_payload = {
                    'product_id': product_id, 'store_id': store_id, 
                    'channel_item_id': str(row.get('Target Product Id', '')).strip(), 
                    'channel_sku': marketplace_sku, 
                    'listed_price': calculated_listed_price, # Easync Fiyatı Yerine Yapay Zeka Fiyatı
                    'quantity': qty,
                    'is_active': True,
                    'needs_sync': False
                }
                # Eğer db'de varsa id'sini ekle ki upsert UPDATE yapsın (böylece var olanların isku'su null olarak kalmaz!)
                if lst_key in listing_cache:
                    l_payload['id'] = listing_cache[lst_key]
                new_listings_payload.append(l_payload)

        # --- 5. TOPLU KAYIT: İLİŞKİLER (UPSERT) ---
        chunk_size = 1000
        
        # Sources
        if new_sources_payload:
            sources_with_id = [item for item in new_sources_payload if 'id' in item]
            sources_without_id = [item for item in new_sources_payload if 'id' not in item]
            
            for sublist in [sources_with_id, sources_without_id]:
                if not sublist:
                    continue
                for i in range(0, len(sublist), chunk_size):
                    try:
                        self.client.table('sources').upsert(sublist[i:i+chunk_size]).execute()
                    except Exception as e:
                        print(f"Source upsert hatası (Length: {len(sublist)}): {e}")
                
        # Listings
        if new_listings_payload:
            listings_with_id = [item for item in new_listings_payload if 'id' in item]
            listings_without_id = [item for item in new_listings_payload if 'id' not in item]
            
            for sublist in [listings_with_id, listings_without_id]:
                if not sublist:
                    continue
                for i in range(0, len(sublist), chunk_size):
                    try:
                        self.client.table('listings').upsert(sublist[i:i+chunk_size]).execute()
                    except Exception as e:
                        print(f"Listings upsert hatası (Length: {len(sublist)}): {e}")

        return {"success": len(df), "errors": 0}

    def get_unapproved_drafts(self) -> List[Dict]:
        """Taslak (draft) tablosundaki incelenmeyi bekleyen kayıtları getirir."""
        # Eskiden limit 500'dü, bot tek seferde yüzlerce ASIN toparlayınca arayüz kitlenmiş sanılıyordu. 5000'e çıkarıldı.
        response = self.client.table("draft").select("*").eq("needs_sync", True).order("created_at", desc=True).limit(5000).execute()
        return response.data if response.data else []

    def import_amazon_drafts(self, df) -> Dict:
        """Amazon'dan indirilen listeleri okur ve draft tablosuna işler."""
        import pandas as pd
        import re
        
        success = 0
        errors = 0
        
        payloads = []
        for index, row in df.iterrows():
            val_0 = str(row.get(0, ""))
            # We must only skip the literal header row, not the products themselves
            if "Example line" in val_0 or (index == 0 and "Line number" in val_0):
                continue
                
            asin = str(row.get(1, "")).strip()
            if not asin or asin.lower() == "nan" or len(asin) < 5:
                continue
                
            # Stok ve Availability (Sütun 6 - Index 6)
            availability = str(row.get(6, "")).strip() if pd.notna(row.get(6)) else ""
            
            qty = 3 # Fırsat avcısından (Deals) gelenler varsayılan 3 stoktur
            avail_lower = availability.lower()
            
            if "currently unavailable" in avail_lower or "out of stock" in avail_lower:
                qty = 0
            elif "only" in avail_lower and "left" in avail_lower:
                match_qty = re.search(r'only\s+(\d+)', avail_lower)
                if match_qty:
                    qty = min(int(match_qty.group(1)), 3)
            elif avail_lower == "0":
                qty = 0
            
            # Extra data toparlama
            extra_data = {}
            if pd.notna(row.get(3)): extra_data["Comment"] = str(row.get(3))
            if pd.notna(row.get(4)): extra_data["Priority"] = str(row.get(4))
            if pd.notna(row.get(5)): extra_data["Validation Check"] = str(row.get(5))
            
            # Fiyat
            price_raw = str(row.get(7, ""))
            price_val = None
            if pd.notna(row.get(7)) and price_raw.lower() != "nan":
                match = re.search(r"(\d+\.\d+|\d+)", price_raw.replace(",", ""))
                if match:
                    price_val = float(match.group(1))
                    
            title = str(row.get(8, "")) if pd.notna(row.get(8)) else "Başlık Yok"
            
            payload = {
                "product_id": asin,
                "title": title[:255],
                "price": price_val,
                "stock_quantity": str(qty),
                "delivery_date": availability[:100],
                "needs_sync": True,
                "extra_data": extra_data
            }
            payloads.append(payload)
            
        # 1) Draft Table Upsert
        chunk_size = 500
        for i in range(0, len(payloads), chunk_size):
            chunk = payloads[i:i+chunk_size]
            try:
                self.client.table('draft').upsert(chunk, on_conflict="product_id").execute()
                success += len(chunk)
            except Exception as e:
                print(f"Draft upsert hatası: {e}")
                errors += len(chunk)
                
        # 2) Live Product Update (Sources & Listings)
        print("Büyük Amazon datası ana tablolara (sources/listings) entegre ediliyor...")
        # Get mapping of ASIN -> product_id
        valid_asins = [p["product_id"] for p in payloads if p["price"] is not None]
        if not valid_asins:
            return {"success": success, "errors": errors}
            
        prod_res = []
        for i in range(0, len(valid_asins), chunk_size):
            chunk = valid_asins[i:i+chunk_size]
            res = self.client.table("core_products").select("id, asin").in_("asin", chunk).execute()
            if res.data:
                prod_res.extend(res.data)
                
        asin_to_pid = {p["asin"]: p["id"] for p in prod_res}
        
        # Get Supplier ID for Amazon US
        sup_data = self.client.table("suppliers").select("id").eq("name", "Amazon US").limit(1).execute()
        supplier_id = sup_data.data[0]["id"] if sup_data.data else None
        
        source_updates = []
        listing_updates = [] # To set needs_sync=True
        
        for p in payloads:
            if p["price"] is None:
                continue
            asin = p["product_id"]
            pid = asin_to_pid.get(asin)
            if not pid or not supplier_id:
                continue
            
            # Update Sources Table (Base Cost)
            source_updates.append({
                "product_id": pid,
                "supplier_id": supplier_id,
                "base_cost": p["price"],
                "source_code": asin
            })
            
            # Queue listing for Push to eBay (Needs Sync)
            listing_updates.append(pid)
            
        if source_updates:
            for i in range(0, len(source_updates), chunk_size):
                chunk = source_updates[i:i+chunk_size]
                self.client.table("sources").upsert(chunk, on_conflict="product_id, supplier_id").execute()
                print(f"- {len(chunk)} adet ana kaynak fiyati guncellendi.")
                
        if listing_updates:
            # We must fetch the actual listing IDs to update them
            list_res = []
            for i in range(0, len(listing_updates), chunk_size):
                chunk = listing_updates[i:i+chunk_size]
                res = self.client.table("listings").select("id, product_id").in_("product_id", chunk).execute()
                if res.data:
                    list_res.extend(res.data)
            
            # Apply Quantity and Needs_Sync
            actual_listing_updates = []
            # Make a fast lookup for qty
            asin_qty_map = {p["product_id"]: int(p["stock_quantity"]) for p in payloads}
            pid_qty_map = {asin_to_pid[asin]: qty for asin, qty in asin_qty_map.items() if asin in asin_to_pid}
            
            for l in list_res:
                pid = l["product_id"]
                new_qty = pid_qty_map.get(pid, 0)
                actual_listing_updates.append({
                    "id": l["id"],
                    "quantity": new_qty,
                    "needs_sync": True
                })
                
            for i in range(0, len(actual_listing_updates), chunk_size):
                chunk = actual_listing_updates[i:i+chunk_size]
                self.client.table("listings").upsert(chunk).execute()
                print(f"- {len(chunk)} adet eBay listelemesi Senkron Kuyruguna eklendi.")

        return {"success": success, "errors": errors}

    def purge_amazon_discard_list(self, df):
        """
        Parses Amazon's error/discard list (which contains duplicates or inactive ASINs),
        extracts the ASINs (usually found in the B column / index 1), and completely removes 
        them from the core_products table to keep the export list pristine.
        """
        success = 0
        errors = 0
        
        # In Amazon's error reports, ASIN is typically in the second column (index 1) after 13 rows of headers.
        # But to be safe if the user uploads a raw list of ASINs, we'll scan the first few columns.
        # CRITICAL FIX: Amazon's report includes ALL ASINs, but only marks bad ones with "Failed" in the Status column.
        import pandas as pd
        asins_to_delete = []
        for index, row in df.iterrows():
            row_vals = [str(x).lower() for x in row.values if pd.notna(x)]
            row_str = " ".join(row_vals)
            
            # Sadece satırda 'failed' veya 'error' ibaresi varsa silinecek ASIN arar
            # (Bu sayede hatasız olan, başarılı eklenen ASIN'leri çöpe atmaz)
            if "failed" in row_str or "error" in row_str:
                found_asin = None
                for col_idx in range(min(5, len(row))):
                    val = str(row.get(col_idx, "")).strip()
                    if len(val) == 10 and " " not in val:
                        found_asin = val
                        break
                
                if found_asin:
                    asins_to_delete.append(found_asin)
        
        # Deduplicate
        asins_to_delete = list(set(asins_to_delete))
        
        # SAFETY MEASURE: Do not allow purging of the entire database by accident
        if len(asins_to_delete) > 500:
            raise ValueError(f"CRITICAL SAFETY LOCK: Trying to delete {len(asins_to_delete)} ASINs at once. This looks like you accidentally uploaded the main export list instead of the Amazon Error/Discard list! Deletion cancelled.")
            
        # Batch delete from core_products
        chunk_size = 200
        for i in range(0, len(asins_to_delete), chunk_size):
            chunk = asins_to_delete[i:i+chunk_size]
            try:
                self.client.table('core_products').delete().in_('asin', chunk).execute()
                success += len(chunk)
            except Exception as e:
                print(f"Purge error: {e}")
                errors += len(chunk)
                
        return {"success": success, "errors": errors, "deleted_asins": asins_to_delete}

    def generate_amazon_export_file(self, output_path):
        """
        Gathers all 10-character ASINs from core_products and formats them into
        the Amazon bulk upload template.
        """
        import openpyxl
        import os
        
        asins = []
        offset = 0
        chunk = 1000
        while True:
            res = self.client.table("core_products").select("asin").not_.is_("asin", "null").range(offset, offset + chunk - 1).execute()
            data = res.data
            chunk_asins = [str(d['asin']).strip() for d in data if d.get('asin') and len(str(d['asin']).strip()) == 10]
            asins.extend(chunk_asins)
            if len(data) < chunk:
                break
            offset += chunk
            
        if asins:
            template_path = os.path.join(os.path.dirname(__file__), "assets", "amazon_template.xlsx")
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            for i, a in enumerate(asins, start=1):
                row_idx = 13 + i
                ws.cell(row=row_idx, column=1, value=i)
                ws.cell(row=row_idx, column=2, value=a)
                
            wb.save(output_path)
            print(f"✅ Amazon listesi oluşturuldu: {output_path} ({len(asins)} ASIN)")
            return True
        else:
            print("❌ Geçerli 10 karakterli ASIN bulunamadı!")
            return False

    def add_product_document(self, product_id: str, document_type: str, document_url: str, listing_id: str = None, language: str = 'en'):
        """Belirtilen ürüne yeni bir döküman (Örn: SDS) ekler."""
        if not product_id or not document_type or not document_url:
            raise ValueError("product_id, document_type ve document_url zorunludur.")
            
        data = {
            "product_id": product_id,
            "document_type": document_type,
            "document_url": document_url,
            "language": language
        }
        
        if listing_id:
            data["listing_id"] = listing_id
            
        return self.client.table("product_documents").insert(data).execute()

    def get_product_documents(self, product_id: str, document_type: str = None) -> List[Dict]:
        """Belirtilen ürünün belgelerini getirir. İstenirse filtreleme (Örn: 'SDS') uygulanabilir."""
        query = self.client.table("product_documents").select("*").eq("product_id", product_id)
        if document_type:
            query = query.eq("document_type", document_type)
            
        res = query.execute()
        return res.data if res.data else []

db = DatabaseManager()